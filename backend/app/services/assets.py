import csv
import io
import uuid
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select, or_, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.asset import Asset
from app.schemas.assets import AssetCreateSchema, AssetUpdateSchema

_COLUMN_MAP = {
    "инв. номер": "inventory_number",
    "инвентарный номер": "inventory_number",
    "inventory_number": "inventory_number",
    "название": "name",
    "наименование": "name",
    "name": "name",
    "баркод": "barcode",
    "штрихкод": "barcode",
    "barcode": "barcode",
    "описание": "description",
    "description": "description",
    "первоначальная стоимость": "initial_cost",
    "стоимость": "initial_cost",
    "initial_cost": "initial_cost",
    "1c id": "one_c_id",
    "1с ид": "one_c_id",
    "one_c_id": "one_c_id",
}


async def list_assets(
    db: AsyncSession,
    page: int = 1,
    limit: int = 50,
    search: str | None = None,
    status: str | None = None,
) -> tuple[list[Asset], int]:
    q = select(Asset).options(selectinload(Asset.photos))
    if search:
        q = q.where(or_(
            Asset.name.ilike(f"%{search}%"),
            Asset.inventory_number.ilike(f"%{search}%"),
            Asset.barcode.ilike(f"%{search}%"),
        ))
    if status:
        q = q.where(Asset.status == status)

    count_q = select(func.count()).select_from(q.subquery())
    total = (await db.scalar(count_q)) or 0

    q = q.order_by(Asset.inventory_number).offset((page - 1) * limit).limit(limit)
    result = await db.execute(q)
    return list(result.scalars().all()), total


async def get_asset(db: AsyncSession, asset_id: str) -> Asset | None:
    try:
        uid = uuid.UUID(asset_id)
    except ValueError:
        return None
    result = await db.execute(
        select(Asset).where(Asset.id == uid).options(selectinload(Asset.photos))
    )
    return result.scalar_one_or_none()


async def get_asset_by_barcode(db: AsyncSession, barcode: str) -> Asset | None:
    result = await db.execute(
        select(Asset).where(Asset.barcode == barcode).options(selectinload(Asset.photos))
    )
    return result.scalar_one_or_none()


async def create_asset(db: AsyncSession, data: AssetCreateSchema) -> Asset:
    asset = Asset(**data.model_dump())
    db.add(asset)
    await db.commit()
    await db.refresh(asset, ["photos"])
    return asset


async def update_asset(db: AsyncSession, asset: Asset, data: AssetUpdateSchema) -> Asset:
    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(asset, key, value)
    await db.commit()
    await db.refresh(asset, ["photos"])
    return asset


async def delete_asset(db: AsyncSession, asset: Asset) -> None:
    await db.delete(asset)
    await db.commit()


async def import_assets_from_file(db: AsyncSession, content: bytes, filename: str) -> dict:
    fname = (filename or "").lower()
    if fname.endswith(".xlsx") or fname.endswith(".xls"):
        rows = _parse_excel(content)
    elif fname.endswith(".csv"):
        rows = _parse_csv(content)
    else:
        raise ValueError("Неподдерживаемый формат. Используйте .xlsx или .csv")

    existing = set(
        r[0] for r in (await db.execute(select(Asset.inventory_number))).all()
    )

    created = 0
    skipped = 0
    errors: list[dict] = []

    for i, row in enumerate(rows, start=2):
        inv_num = str(row.get("inventory_number") or "").strip()
        name = str(row.get("name") or "").strip()

        if not inv_num:
            errors.append({"row": i, "message": "Пустой инвентарный номер"})
            continue
        if not name:
            errors.append({"row": i, "message": "Отсутствует название"})
            continue
        if inv_num in existing:
            skipped += 1
            continue

        initial_cost = None
        raw_cost = row.get("initial_cost")
        if raw_cost is not None and str(raw_cost).strip():
            try:
                initial_cost = Decimal(str(raw_cost).replace(",", ".").replace(" ", "").replace("\xa0", ""))
            except InvalidOperation:
                pass

        try:
            db.add(Asset(
                inventory_number=inv_num,
                name=name,
                barcode=str(row["barcode"]).strip() or None if row.get("barcode") else None,
                description=str(row["description"]).strip() or None if row.get("description") else None,
                initial_cost=initial_cost,
                one_c_id=str(row["one_c_id"]).strip() or None if row.get("one_c_id") else None,
            ))
            existing.add(inv_num)
            created += 1
        except Exception as exc:
            errors.append({"row": i, "message": str(exc)})

    if created:
        await db.commit()

    return {"created": created, "skipped": skipped, "errors": errors}


def generate_import_template() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Активы"

    headers = ["Инв. номер", "Название", "Баркод", "Описание", "Первоначальная стоимость", "1C ID"]
    ws.append(headers)
    ws.append(["INV-001", "Компьютер Dell", "1234567890", "Офис 101", "50000.00", ""])

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col, width in zip("ABCDEF", [16, 32, 15, 32, 22, 16]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_excel(content: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    raw_headers = next(rows_iter, [])
    fields = [_COLUMN_MAP.get(str(h).strip().lower(), "") if h is not None else "" for h in raw_headers]

    result = []
    for row in rows_iter:
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        d = {field: value for field, value in zip(fields, row) if field}
        if d:
            result.append(d)
    wb.close()
    return result


def _parse_csv(content: bytes) -> list[dict]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = content.decode("cp1251")

    # Auto-detect delimiter
    delimiter = ";" if text.count(";") >= text.count(",") else ","

    result = []
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    for row in reader:
        d = {}
        for key, value in row.items():
            field = _COLUMN_MAP.get((key or "").strip().lower(), "")
            if field and value and value.strip():
                d[field] = value.strip()
        if d:
            result.append(d)
    return result
